#!/usr/bin/env python3
"""
ULBC Trust — FileMaker to Salesforce Migration Script
Phase 4: Full data migration

Usage:
    pip install -r requirements.txt
    python migrate.py --step contacts        # Step 1: Contacts
    python migrate.py --step education       # Step 2: Education History
    python migrate.py --step opportunities   # Step 3: Opportunities
    python migrate.py --step crews           # Step 4: Crew History
    python migrate.py --step events          # Step 5: Event Campaigns
    python migrate.py --step attendance      # Step 6: Event Attendance
    python migrate.py --step notes           # Step 7: ContentNotes
    python migrate.py --step partners        # Step 8: Partner Relationships
    python migrate.py --step all             # Run all steps in order

Requires: sf CLI authenticated to org alias 'ulbc'
"""

import argparse
import base64
import csv
import json
import logging
import os
import subprocess
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

import openpyxl
from simple_salesforce import Salesforce, SalesforceError

# ── Configuration ────────────────────────────────────────────────────────

SOURCE_DIR = Path(__file__).parent / "source"
LOG_DIR = Path(__file__).parent / "migration_log"
BATCH_SIZE = 200

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / "migration.log"),
    ],
)
log = logging.getLogger("migrate")


# ── Salesforce Connection ────────────────────────────────────────────────

def connect_sf() -> Salesforce:
    """Connect to Salesforce using sf CLI session."""
    result = subprocess.run(
        ["sf", "org", "display", "--target-org", "ulbc", "--json"],
        capture_output=True, text=True,
    )
    org_info = json.loads(result.stdout)["result"]
    return Salesforce(
        instance_url=org_info["instanceUrl"],
        session_id=org_info["accessToken"],
    )


# ── Helpers ──────────────────────────────────────────────────────────────

def clean_str(val: Any) -> Optional[str]:
    """Clean a cell value: strip, replace FileMaker encoding artefacts."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None"):
        return None
    # Fix FileMaker/Latin-1 encoding artefacts
    s = s.replace("\u00ca", "").replace("\u00cd", "'").replace("\u00d5", "'")
    s = s.replace("\x0b", "\n")  # vertical tab → newline
    return s


def parse_date_dmy(val: Any) -> Optional[str]:
    """Parse DD/MM/YYYY string to YYYY-MM-DD."""
    s = clean_str(val)
    if not s:
        return None
    try:
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None
    except Exception:
        return None


def parse_date_xl(val: Any) -> Optional[str]:
    """Parse Excel datetime object to YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return parse_date_dmy(val)


def safe_float(val: Any) -> Optional[float]:
    """Parse a numeric value, stripping currency symbols."""
    s = clean_str(val)
    if not s:
        return None
    s = s.replace("$", "").replace("£", "").replace(",", "").replace("Cn", "")
    try:
        return float(s)
    except ValueError:
        return None


def is_valid_trust_id(tid: Any) -> bool:
    """Check if TrustId is a valid numeric ID."""
    s = clean_str(tid)
    if not s:
        return False
    if s in ("#N/A", "---", "0", "#REF!"):
        return False
    try:
        int(s)
        return True
    except ValueError:
        return False


def batch_insert(sf: Salesforce, sobject: str, records: list[dict],
                 log_file: str) -> tuple[int, int]:
    """Insert records in batches, log failures."""
    success = 0
    failed = 0
    failed_records = []
    sobj = getattr(sf.bulk, sobject)

    for i in range(0, len(records), BATCH_SIZE):
        batch = records[i:i + BATCH_SIZE]
        try:
            results = sobj.insert(batch)
            for j, r in enumerate(results):
                if r.get("success"):
                    success += 1
                else:
                    failed += 1
                    failed_records.append({
                        "index": i + j,
                        "errors": r.get("errors", []),
                        "record": batch[j],
                    })
        except SalesforceError as e:
            log.error(f"Batch {i}-{i+len(batch)} failed: {e}")
            failed += len(batch)
            for j, rec in enumerate(batch):
                failed_records.append({
                    "index": i + j,
                    "errors": [str(e)],
                    "record": rec,
                })

    if failed_records:
        with open(LOG_DIR / log_file, "w") as f:
            json.dump(failed_records, f, indent=2, default=str)
        log.warning(f"  {failed} failures logged to {log_file}")

    return success, failed


def batch_upsert(sf: Salesforce, sobject: str, external_id: str,
                 records: list[dict], log_file: str) -> tuple[int, int]:
    """Upsert records in batches using an external ID field."""
    success = 0
    failed = 0
    failed_records = []
    sobj = getattr(sf.bulk, sobject)

    for i in range(0, len(records), BATCH_SIZE):
        batch = records[i:i + BATCH_SIZE]
        try:
            results = sobj.upsert(batch, external_id)
            for j, r in enumerate(results):
                if r.get("success"):
                    success += 1
                else:
                    failed += 1
                    failed_records.append({
                        "index": i + j,
                        "errors": r.get("errors", []),
                        "record": batch[j],
                    })
        except SalesforceError as e:
            log.error(f"Batch {i}-{i+len(batch)} failed: {e}")
            failed += len(batch)

    if failed_records:
        with open(LOG_DIR / log_file, "w") as f:
            json.dump(failed_records, f, indent=2, default=str)
        log.warning(f"  {failed} failures logged to {log_file}")

    return success, failed


# ── Step 1: Contacts ────────────────────────────────────────────────────

MEM_TYPE_MAP = {
    "A": "Alumni",
    "S": "Other",
    "O": "Other",
    "P": "Other",
    "D": "Other",
}

SEX_MAP = {
    "M": "Male",
    "W": "Female",
    "w": "Female",
}

# Sub type code → Legacy Membership label
SUB_TYPE_MAP = {
    "P": "Patron",
    "D": "Deceased Donor",
    "X": "Excluded",
    "x": "Excluded",
}

# Honorary Patrons — TrustIDs from patron analysis spreadsheet (Mar 2026).
# These are recognised for service, not giving.
HONORARY_PATRON_TIDS = {"44", "151", "293", "974", "1028", "1044", "1046", "1147"}


def load_contacts() -> list[dict]:
    """Load and transform Main Contact.xlsx → Contact records."""
    wb = openpyxl.load_workbook(SOURCE_DIR / "Main Contact.xlsx", read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    wb.close()

    records = []
    for row in rows[1:]:
        d = dict(zip(header, row))
        tid = clean_str(d.get("TrustID"))
        if not tid:
            continue

        mem_type = clean_str(d.get("Mem Type")) or ""
        sub_code = clean_str(d.get("Member Sub type Code")) or ""
        desc_parts = [f"FileMaker Mem Type: {mem_type}"]
        if sub_code:
            desc_parts.append(f"Sub Code: {sub_code}")

        # Patron Status and Legacy Membership from sub type code
        legacy_membership = SUB_TYPE_MAP.get(sub_code)
        patron_status = "Non-Patron"
        if sub_code == "P":
            if tid in HONORARY_PATRON_TIDS:
                patron_status = "Honorary Patron"
                legacy_membership = "Patron, Honorary"
            else:
                patron_status = "Patron"

        # email opt in: Y → not opted out, N → opted out, blank → not opted out
        email_opt = clean_str(d.get("email opt in"))
        has_opted_out = email_opt == "N"

        sex = clean_str(d.get("Sex"))
        gender = SEX_MAP.get(sex) if sex else None

        # Address: comma-separated, split into street lines
        address = clean_str(d.get("Address"))
        mailing_street = address.replace(",", "\n").strip() if address else None

        postcode = clean_str(d.get("Post Code"))
        if postcode:
            postcode = postcode.replace("\u00ca", "").strip()

        rec = {
            "ULBC_Trust_ID__c": tid,
            "Salutation": clean_str(d.get("Title")),
            "FirstName": clean_str(d.get("First Name")),
            "LastName": clean_str(d.get("Last Name")),
            "Birthdate": parse_date_xl(d.get("Birthdate")),
            "ULBC_Birth_Name__c": clean_str(d.get("Birth Name")),
            "Email": clean_str(d.get("email")),
            "ULBC_Secondary_Email__c": clean_str(d.get("email2")),
            "HomePhone": clean_str(d.get("Home Phone")),
            "MobilePhone": clean_str(d.get("Mobile Phone")),
            "Phone": clean_str(d.get("Workph")),
            "MailingStreet": mailing_street,
            "MailingCity": clean_str(d.get("Town")),
            "MailingState": clean_str(d.get("County")),
            "MailingPostalCode": postcode,
            "MailingCountry": clean_str(d.get("Country")),
            "ULBC_Gender__c": gender,
            "ULBC_Primary_Contact_Type__c": MEM_TYPE_MAP.get(mem_type, "Other"),
            "HasOptedOutOfEmail": has_opted_out,
            "ULBC_Gift_Aid_Declaration_Date__c": parse_date_xl(d.get("Gift Aid decl date")),
            "ULBC_Gift_Aid_Declaration_Source__c": clean_str(d.get("Gift aid decl source")),
            "ULBC_Gift_Aid_Valid_From__c": parse_date_xl(d.get("Gift Aid from date")),
            "ULBC_Gift_Aid_Valid_To__c": parse_date_xl(d.get("Gift Aid to date")),
            "ULBC_Is_Volunteer__c": bool(clean_str(d.get("Volunteer"))),
            "ULBC_Profession__c": clean_str(d.get("profession")),
            "ULBC_Other_Interests__c": clean_str(d.get("Other interests")),
            "ULBC_Special_Skills__c": clean_str(d.get("Special skills")),
            "ULBC_Follow_Up_Notes__c": clean_str(d.get("Follow up notes")),
            "ULBC_Year_Group__c": clean_str(d.get("Year Group")),
            "ULBC_School__c": clean_str(d.get("School")),
            "ULBC_Subscription_Start_Date__c": parse_date_xl(d.get("Date of first sub")),
            "ULBC_Subscription_Last_Change__c": parse_date_xl(d.get("Date last sub change")),
            "ULBC_Subscription_Lapsed_Date__c": parse_date_xl(d.get("sub lapsed date")),
            "ULBC_Patron_Status__c": patron_status,
            "ULBC_GDPR_Legal_Basis__c": "Legitimate Interests",
            "ULBC_Acquisition_Channel__c": "Alumni",
            "Description": ", ".join(desc_parts),
        }

        if legacy_membership:
            rec["ULBC_Legacy_Membership__c"] = legacy_membership

        # Deceased flag
        if mem_type == "D":
            rec["npsp__Deceased__c"] = True

        # Remove None values
        rec = {k: v for k, v in rec.items() if v is not None}
        records.append(rec)

    return records


def step_contacts(sf: Salesforce) -> None:
    """Step 1: Upsert contacts from Main Contact.xlsx."""
    log.info("Step 1: Contacts")
    records = load_contacts()
    log.info(f"  Loaded {len(records)} contacts from xlsx")

    success, failed = batch_upsert(
        sf, "Contact", "ULBC_Trust_ID__c", records, "contacts_failed.json"
    )
    log.info(f"  Contacts: {success} success, {failed} failed")


# ── Step 2: Education History ────────────────────────────────────────────

DEGREE_MAP = {
    "BA": "Undergraduate",
    "BSc": "Undergraduate",
    "BEng": "Undergraduate",
    "LLB": "Undergraduate",
    "MBBS": "Undergraduate",
    "MA": "Masters",
    "MSc": "Masters",
    "MBA": "Masters",
    "MPhil": "Masters",
    "PhD": "PhD",
    "DPhil": "PhD",
}


def step_education(sf: Salesforce) -> None:
    """Step 2: Create Education History from Main Contact.xlsx."""
    log.info("Step 2: Education History")

    # Build TrustID → Contact Id map
    contacts = sf.query_all(
        "SELECT Id, ULBC_Trust_ID__c FROM Contact WHERE ULBC_Trust_ID__c != null"
    )
    tid_map = {r["ULBC_Trust_ID__c"]: r["Id"] for r in contacts["records"]}
    log.info(f"  {len(tid_map)} contacts in org with TrustID")

    wb = openpyxl.load_workbook(SOURCE_DIR / "Main Contact.xlsx", read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    wb.close()

    records = []
    skipped = 0
    for row in rows[1:]:
        d = dict(zip(header, row))
        tid = clean_str(d.get("TrustID"))

        college = clean_str(d.get("College ID"))
        degree = clean_str(d.get("Degree"))
        course = clean_str(d.get("Course"))
        first_year = d.get("FirstYearatUL")
        last_year = d.get("LastYearatUL")

        # Skip if no education data
        if not any([college, degree, course, first_year, last_year]):
            continue

        contact_id = tid_map.get(tid)
        if not contact_id:
            skipped += 1
            continue

        # Map degree to picklist
        degree_type = None
        if degree:
            degree_str = degree.strip()
            degree_type = DEGREE_MAP.get(degree_str, "Other")

        rec = {"ULBC_Contact__c": contact_id}
        if college:
            rec["ULBC_Institution__c"] = college
        if degree_type:
            rec["ULBC_Degree_Type__c"] = degree_type
        if course:
            rec["ULBC_Subject__c"] = clean_str(course)
        if first_year is not None:
            try:
                rec["ULBC_Start_Year__c"] = int(first_year)
            except (ValueError, TypeError):
                pass
        if last_year is not None:
            try:
                rec["ULBC_End_Year__c"] = int(last_year)
            except (ValueError, TypeError):
                pass

        records.append(rec)

    log.info(f"  {len(records)} education records to create, {skipped} skipped (no contact)")

    success, failed = batch_insert(
        sf, "ULBC_Education_History__c", records, "education_failed.json"
    )
    log.info(f"  Education History: {success} success, {failed} failed")


# ── Step 3: Opportunities ───────────────────────────────────────────────

FUND_TYPE_MAP = {
    "S": "Unrestricted",
    "DM": "Unrestricted",
    "DN": "Restricted",
    "DS": "Unrestricted",
}

GIFT_TYPE_MAP = {
    "S": "Regular",
    "DM": "One-off",
}


def step_opportunities(sf: Salesforce) -> None:
    """Step 3: Create Opportunities from Accounts.csv."""
    log.info("Step 3: Opportunities")

    # Build TrustID → Contact Id + AccountId map
    contacts = sf.query_all(
        "SELECT Id, AccountId, ULBC_Trust_ID__c FROM Contact "
        "WHERE ULBC_Trust_ID__c != null"
    )
    tid_map = {}
    for r in contacts["records"]:
        tid_map[r["ULBC_Trust_ID__c"]] = {
            "ContactId": r["Id"],
            "AccountId": r["AccountId"],
        }
    log.info(f"  {len(tid_map)} contacts mapped")

    records = []
    skipped = []

    with open(SOURCE_DIR / "Accounts.csv", "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            tid = row.get("TrustId", "").strip()
            if not is_valid_trust_id(tid):
                skipped.append({"row": i + 2, "reason": "invalid TrustId", "tid": tid})
                continue

            contact_info = tid_map.get(tid)
            if not contact_info:
                skipped.append({"row": i + 2, "reason": "TrustId not found", "tid": tid})
                continue

            account_code = row.get("Account code", "").strip().upper()
            mu = row.get("Account M U", "").strip().upper()
            desc = clean_str(row.get("Description")) or ""
            amount = safe_float(row.get("Amount"))
            close_date = parse_date_dmy(row.get("Date"))

            if not close_date:
                skipped.append({"row": i + 2, "reason": "invalid date", "date": row.get("Date")})
                continue

            if amount is None:
                skipped.append({"row": i + 2, "reason": "invalid amount", "amount": row.get("Amount")})
                continue

            # Build Opportunity name
            opp_name = desc[:120] if desc else f"FileMaker Import #{row.get('Seq No', i)}"

            rec = {
                "AccountId": contact_info["AccountId"],
                "npsp__Primary_Contact__c": contact_info["ContactId"],
                "Name": opp_name,
                "StageName": "Closed Won",
                "CloseDate": close_date,
                "Amount": abs(amount),
                "Description": f"Account code: {account_code} | M/U: {mu} | {desc}",
            }

            fund_type = FUND_TYPE_MAP.get(account_code)
            if fund_type:
                rec["ULBC_Fund_Type__c"] = fund_type

            gift_type = GIFT_TYPE_MAP.get(account_code)
            if gift_type:
                rec["ULBC_Gift_Type__c"] = gift_type

            # Gift Aid
            ga = clean_str(row.get("Gift Aid Eligible"))
            if ga:
                rec["ULBC_Gift_Aid_Eligible__c"] = True

            ga_date = parse_date_dmy(row.get("Gift aid Tax claim date"))
            if ga_date:
                rec["ULBC_Gift_Aid_Claimed_Date__c"] = ga_date

            records.append(rec)

    log.info(f"  {len(records)} opportunities to create, {len(skipped)} skipped")

    if skipped:
        with open(LOG_DIR / "opportunities_skipped.json", "w") as f:
            json.dump(skipped, f, indent=2, default=str)
        log.info(f"  Skipped records logged to opportunities_skipped.json")

    success, failed = batch_insert(
        sf, "Opportunity", records, "opportunities_failed.json"
    )
    log.info(f"  Opportunities: {success} success, {failed} failed")


# ── Step 4: Crew History ─────────────────────────────────────────────────

def step_crews(sf: Salesforce) -> None:
    """Step 4: Create Crew History from Crew Code + Crew Name CSVs."""
    log.info("Step 4: Crew History")

    # Build TrustID → Contact Id map
    contacts = sf.query_all(
        "SELECT Id, ULBC_Trust_ID__c FROM Contact WHERE ULBC_Trust_ID__c != null"
    )
    tid_map = {r["ULBC_Trust_ID__c"]: r["Id"] for r in contacts["records"]}

    # Load crew name lookup
    crew_names = {}
    with open(SOURCE_DIR / "Crew Name.csv", "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row.get("Crew Code", "").strip()
            if code:
                crew_names[code] = {
                    "year": row.get("Year", "").strip(),
                    "short_name": clean_str(row.get("Short Name")),
                    "crew_name": clean_str(row.get("Crew Name")),
                    "regatta": clean_str(row.get("Regatta")),
                    "result": clean_str(row.get("Race Position")),
                }
    log.info(f"  {len(crew_names)} crew names loaded")

    # Load crew code (person → crew) and join with crew names
    records = []
    skipped = 0
    with open(SOURCE_DIR / "Crew Code.csv", "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        for row in reader:
            tid = row.get("TrustId", "").strip()
            code = row.get("Crew Code", "").strip()
            position = row.get("Rowing Position", "").strip()

            contact_id = tid_map.get(tid)
            if not contact_id:
                skipped += 1
                continue

            rec = {
                "ULBC_Contact__c": contact_id,
                "ULBC_Crew_Code__c": code,
            }

            # Position
            if position:
                try:
                    rec["ULBC_Rowing_Position__c"] = int(position)
                except ValueError:
                    pass

            # Join with crew name details
            cn = crew_names.get(code)
            if cn:
                if cn["year"]:
                    try:
                        rec["ULBC_Year__c"] = int(cn["year"])
                    except ValueError:
                        pass
                if cn["crew_name"]:
                    rec["ULBC_Crew_Name__c"] = cn["crew_name"]
                if cn["regatta"]:
                    rec["ULBC_Regatta__c"] = cn["regatta"]
                if cn["short_name"]:
                    rec["ULBC_Event__c"] = cn["short_name"]
                if cn["result"]:
                    rec["ULBC_Result__c"] = cn["result"]

            records.append(rec)

    log.info(f"  {len(records)} crew records to create, {skipped} skipped (no contact)")

    success, failed = batch_insert(
        sf, "ULBC_Crew_History__c", records, "crews_failed.json"
    )
    log.info(f"  Crew History: {success} success, {failed} failed")


# ── Step 5: Event Campaigns ─────────────────────────────────────────────

def step_events(sf: Salesforce) -> None:
    """Step 5: Create event Campaigns from Event Name.csv."""
    log.info("Step 5: Event Campaigns")

    with open(SOURCE_DIR / "Event Name.csv", "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        records = []
        for row in reader:
            event_id = clean_str(row.get("Event Id"))
            event_name = clean_str(row.get("Event Name"))
            year = clean_str(row.get("Year"))

            if not event_id or not event_name:
                continue

            name = f"{event_name} {year}" if year else event_name
            start_date = f"{year}-01-01" if year else None

            rec = {
                "Name": name,
                "Type": "Event",
                "Status": "Completed",
                "IsActive": False,
                "Description": f"FileMaker Event ID: {event_id}",
            }
            if start_date:
                rec["StartDate"] = start_date

            records.append(rec)

    log.info(f"  {len(records)} event campaigns to create")

    # Insert one by one to track Event ID → Campaign Id mapping
    event_map = {}
    success = 0
    failed = 0
    for i, rec in enumerate(records):
        try:
            result = sf.Campaign.create(rec)
            if result.get("success"):
                # Extract event_id from description
                desc = rec["Description"]
                eid = desc.replace("FileMaker Event ID: ", "")
                event_map[eid] = result["id"]
                success += 1
            else:
                failed += 1
                log.error(f"  Failed to create campaign: {rec['Name']}: {result}")
        except SalesforceError as e:
            failed += 1
            log.error(f"  Failed to create campaign: {rec['Name']}: {e}")

    # Save event map for attendance step
    with open(LOG_DIR / "event_campaign_map.json", "w") as f:
        json.dump(event_map, f, indent=2)

    log.info(f"  Events: {success} success, {failed} failed")
    log.info(f"  Event map saved to event_campaign_map.json")


# ── Step 6: Event Attendance ─────────────────────────────────────────────

def step_attendance(sf: Salesforce) -> None:
    """Step 6: Create CampaignMembers from Event attendance.csv."""
    log.info("Step 6: Event Attendance")

    # Load event map
    map_path = LOG_DIR / "event_campaign_map.json"
    if not map_path.exists():
        log.error("  event_campaign_map.json not found — run step 'events' first")
        return
    with open(map_path) as f:
        event_map = json.load(f)
    log.info(f"  {len(event_map)} events in map")

    # Build TrustID → Contact Id map
    contacts = sf.query_all(
        "SELECT Id, ULBC_Trust_ID__c FROM Contact WHERE ULBC_Trust_ID__c != null"
    )
    tid_map = {r["ULBC_Trust_ID__c"]: r["Id"] for r in contacts["records"]}

    # Create 'Attended' status on each campaign that doesn't have it
    for event_id, campaign_id in event_map.items():
        existing = sf.query(
            f"SELECT Id FROM CampaignMemberStatus "
            f"WHERE CampaignId = '{campaign_id}' AND Label = 'Attended'"
        )
        if existing["totalSize"] == 0:
            try:
                sf.CampaignMemberStatus.create({
                    "CampaignId": campaign_id,
                    "Label": "Attended",
                    "SortOrder": 3,
                    "IsDefault": False,
                    "HasResponded": True,
                })
            except SalesforceError as e:
                log.warning(f"  Could not create Attended status on {event_id}: {e}")

    # Parse attendance data — all on one line, need to handle carefully
    with open(SOURCE_DIR / "Event attendance.csv", "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        records = []
        skipped = []
        seen = set()

        for row in reader:
            event_id = clean_str(row.get("Event ID"))
            tid = clean_str(row.get("TrustId"))

            if not event_id or not tid:
                continue

            campaign_id = event_map.get(event_id)
            if not campaign_id:
                skipped.append({"event_id": event_id, "tid": tid, "reason": "event not found"})
                continue

            contact_id = tid_map.get(tid)
            if not contact_id:
                skipped.append({"event_id": event_id, "tid": tid, "reason": "contact not found"})
                continue

            # Deduplicate — same person at same event
            key = f"{campaign_id}:{contact_id}"
            if key in seen:
                continue
            seen.add(key)

            records.append({
                "CampaignId": campaign_id,
                "ContactId": contact_id,
                "Status": "Attended",
            })

    log.info(f"  {len(records)} attendance records (deduplicated), {len(skipped)} skipped")

    if skipped:
        with open(LOG_DIR / "attendance_skipped.json", "w") as f:
            json.dump(skipped, f, indent=2)

    success, failed = batch_insert(
        sf, "CampaignMember", records, "attendance_failed.json"
    )
    log.info(f"  Attendance: {success} success, {failed} failed")


# ── Step 7: Notes ────────────────────────────────────────────────────────

def step_notes(sf: Salesforce) -> None:
    """Step 7: Create Notes (standard Note object) from Notes.csv."""
    log.info("Step 7: Notes")

    # Build TrustID → Contact Id map
    contacts = sf.query_all(
        "SELECT Id, ULBC_Trust_ID__c FROM Contact WHERE ULBC_Trust_ID__c != null"
    )
    tid_map = {r["ULBC_Trust_ID__c"]: r["Id"] for r in contacts["records"]}

    records = []
    skipped = 0

    with open(SOURCE_DIR / "Notes.csv", "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        for row in reader:
            tid = clean_str(row.get("TrustId"))
            subject = clean_str(row.get("Note subject")) or "Note"
            body = clean_str(row.get("Note")) or ""

            contact_id = tid_map.get(tid)
            if not contact_id:
                skipped += 1
                continue

            records.append({
                "ParentId": contact_id,
                "Title": subject[:255],
                "Body": body[:32000],
                "IsPrivate": False,
            })

    log.info(f"  {len(records)} notes to create, {skipped} skipped (no contact)")

    success, failed = batch_insert(
        sf, "Note", records, "notes_failed.json"
    )
    log.info(f"  Notes: {success} success, {failed} failed")


# ── Step 8: Partner Relationships ────────────────────────────────────────

def step_partners(sf: Salesforce) -> None:
    """Step 8: Create NPSP Relationships for partners."""
    log.info("Step 8: Partner Relationships")

    # Build TrustID → Contact Id map
    contacts = sf.query_all(
        "SELECT Id, ULBC_Trust_ID__c FROM Contact WHERE ULBC_Trust_ID__c != null"
    )
    tid_map = {r["ULBC_Trust_ID__c"]: r["Id"] for r in contacts["records"]}

    wb = openpyxl.load_workbook(SOURCE_DIR / "Main Contact.xlsx", read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    wb.close()

    success = 0
    failed = 0
    skipped = 0

    for row in rows[1:]:
        d = dict(zip(header, row))
        tid = clean_str(d.get("TrustID"))
        partner_tid = clean_str(d.get("Partners TrustID"))

        if not partner_tid or partner_tid == "0":
            continue

        contact_id = tid_map.get(tid)
        partner_id = tid_map.get(partner_tid)

        if not contact_id or not partner_id:
            skipped += 1
            continue

        try:
            sf.npe4__Relationship__c.create({
                "npe4__Contact__c": contact_id,
                "npe4__RelatedContact__c": partner_id,
                "npe4__Type__c": "Partner",
                "npe4__Status__c": "Current",
            })
            success += 1
        except SalesforceError as e:
            failed += 1
            log.error(f"  Relationship {tid} → {partner_tid} failed: {e}")

    log.info(f"  Relationships: {success} success, {failed} failed, {skipped} skipped")


# ── Step 9: Account Transactions ────────────────────────────────────────

def step_transactions(sf: Salesforce) -> None:
    """Step 9: Load raw transactions from Accounts.csv into Account_Transaction__c."""
    log.info("Step 9: Account Transactions")

    # Build TrustID → Account Id map (TrustID__c on Account_Transaction is a lookup to Account)
    contacts = sf.query_all(
        "SELECT Id, AccountId, ULBC_Trust_ID__c FROM Contact WHERE ULBC_Trust_ID__c != null"
    )
    tid_map = {r["ULBC_Trust_ID__c"]: r["AccountId"] for r in contacts["records"] if r.get("AccountId")}
    log.info(f"  {len(tid_map)} contacts mapped to accounts")

    records = []
    skipped = []

    with open(SOURCE_DIR / "Accounts.csv", "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            tid = row.get("TrustId", "").strip()
            amount = safe_float(row.get("Amount"))
            close_date = parse_date_dmy(row.get("Date"))
            desc = clean_str(row.get("Description")) or ""
            mu = row.get("Account M U", "").strip()
            seq_no = row.get("Seq No", "").strip()
            ac_import = row.get("Ac import no", "").strip()
            currency_amount = clean_str(row.get("Currency amount"))
            ga_eligible = clean_str(row.get("Gift Aid Eligible"))
            ga_claim_date = clean_str(row.get("Gift aid Tax claim date"))

            rec = {
                "Date__c": close_date or "",
                "Description__c": desc,
                "Account_M_U__c": mu,
                "Seq_No__c": seq_no,
                "Ac_import_no__c": ac_import,
                "Gift_Aid_Eligible__c": ga_eligible or "",
                "Gift_aid_Tax_claim_date__c": ga_claim_date or "",
            }

            if amount is not None:
                rec["Amount__c"] = abs(amount)
                rec["Total_Amount__c"] = abs(amount)

            if currency_amount:
                rec["Currency_amount__c"] = currency_amount

            # Link to Contact via TrustID lookup
            if tid and is_valid_trust_id(tid):
                contact_id = tid_map.get(tid)
                if contact_id:
                    rec["TrustID__c"] = contact_id

            records.append(rec)

    log.info(f"  {len(records)} transactions to create")

    success, failed = batch_insert(
        sf, "Account_Transaction__c", records, "transactions_failed.json"
    )
    log.info(f"  Transactions: {success} success, {failed} failed")


# ── Main ─────────────────────────────────────────────────────────────────

STEPS = {
    "contacts": step_contacts,
    "education": step_education,
    "opportunities": step_opportunities,
    "crews": step_crews,
    "events": step_events,
    "attendance": step_attendance,
    "notes": step_notes,
    "partners": step_partners,
    "transactions": step_transactions,
}

ALL_STEPS = [
    "contacts", "education", "opportunities", "crews",
    "events", "attendance", "notes", "partners", "transactions",
]


def main() -> None:
    parser = argparse.ArgumentParser(description="ULBC FileMaker → Salesforce migration")
    parser.add_argument("--step", required=True, choices=list(STEPS.keys()) + ["all"],
                        help="Which migration step to run")
    args = parser.parse_args()

    LOG_DIR.mkdir(exist_ok=True)

    log.info("Connecting to Salesforce...")
    sf = connect_sf()
    log.info(f"Connected to {sf.sf_instance}")

    if args.step == "all":
        for step_name in ALL_STEPS:
            STEPS[step_name](sf)
            log.info("")
    else:
        STEPS[args.step](sf)

    log.info("Done.")


if __name__ == "__main__":
    main()
